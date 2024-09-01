import os
from flask import Flask, send_file
from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl

app = Flask(__name__)

def update_excel_file():
    chrome_options = webdriver.ChromeOptions()
    chrome_options.binary_location = "chrome"
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(options=chrome_options)

    urls = [
        "https://stash.clash.gg/containers/skin-cases",
        "https://stash.clash.gg/stickers/capsule/294/CS20-Sticker-Capsule"
    ]

    html_contents = {}

    for i, url in enumerate(urls):
        driver.get(url)
        html_content = driver.page_source
        html_contents[i] = html_content

    driver.quit()

    cases = [
        "Kilowatt Case",
        "Snakebite Case",
        "Revolution Case",
        "Dreams & Nightmares Case",
        "Clutch Case",
        "Danger Zone Case",
        "Fracture Case",
        "Prisma 2 Case",
        "Prisma Case",
        "CS20 Case",
        "Spectrum 2 Case",
        "Gamma 2 Case",
        "Glove Case",
        "Horizon Case",
        "CS20 Sticker Capsule",
        "Recoil Case",
        "CS:GO Weapon Case 2",
        "Operation Phoenix Weapon Case",
        "Revolver Case",
        "Shadow Case",
        "Chroma 3 Case"
    ]

    obtained_cases = {}

    # Process the first URL (Skin Cases)
    soup = BeautifulSoup(html_contents[0], 'html.parser')
    containers = soup.find_all('div', class_='well result-box nomargin')

    for container in containers:
        title = container.find('h4').text.strip()
        price_container = container.find('div', 'price margin-top-sm')
        price_text = price_container.find('p').text.strip()
        price = price_text.replace('$', '')

        if title in cases and title not in obtained_cases:
            obtained_cases[title] = price

    # Process the second URL (Sticker Capsule)
    soup = BeautifulSoup(html_contents[1], 'html.parser')
    capsules = soup.find_all('div', class_='col-lg-12 text-center col-widen content-header')

    for capsule in capsules:
        title_container = capsule.find('div', class_='inline-middle collapsed-top-margin')
        title = title_container.find('h1').text.strip()
        price_container = capsule.find('a', class_='btn btn-default market-button-item')
        price_text = price_container.text.strip()
        price = price_text.split()[0].replace('$', '')

        if title in cases and title not in obtained_cases:
            obtained_cases[title] = price

    # Load and update the Excel file
    workbook = openpyxl.load_workbook('CS2 Cases.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        case_name = row[0].value  # Column A
        if case_name in obtained_cases:
            row[5].value = float(obtained_cases[case_name])  # Column F

    # Save the changes
    workbook.save('CS2 Cases.xlsx')

@app.route('/')
def download_file():
    # Update the Excel file before sending it
    update_excel_file()
    
    # Define the path to the file
    file_path = 'CS2 Cases.xlsx'

    # Send the file to the user
    return send_file(file_path, as_attachment=True)

if __name__ == "__main__":
    # Determine if the script is running on Heroku
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
